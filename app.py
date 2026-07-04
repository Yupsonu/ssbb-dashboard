"""
SSBB Project Dashboard — Streamlit version
Mirrors the Power BI "SSBB PROJECT DASHBOARD" and auto-refreshes whenever the
source Excel file (SSBB.xlsx) is updated on disk.
"""

import os
import time
import pandas as pd
import numpy as np
import streamlit as st
import plotly.graph_objects as go

# Optional auto-refresh (falls back gracefully if not installed)
try:
    from streamlit_autorefresh import st_autorefresh
    HAS_AUTOREFRESH = True
except ImportError:
    HAS_AUTOREFRESH = False

SHEET_NAME = "SSBB Batch 1&2 Tracking "

# Column positions (1-indexed, matching the source workbook layout)
COL = {
    "Batch": 1,
    "SNo": 2,
    "EmplID": 3,
    "Name": 4,
    "Mobile": 5,
    "Location": 6,
    "Department": 7,
    "ProjectTitle": 8,
    "CTQ": 9,
    "UoM": 10,
    "Baseline": 11,
    "Target": 12,
    "TB_Annually": 13,
    "ARS": 14,
    "ReviewSchedule": 15,
    "D": 16,
    "M": 21,
    "A": 25,
    "I": 30,
    "C": 32,
    "ProjectLink": 39,
}

st.set_page_config(
    page_title="SSBB Project Dashboard",
    page_icon="📊",
    layout="wide",
)


# ----------------------------------------------------------------------
# Data loading
# ----------------------------------------------------------------------
def get_file_signature(path: str):
    """Return (mtime, size) so cache invalidates whenever the file changes."""
    stat = os.stat(path)
    return stat.st_mtime, stat.st_size


@st.cache_data(show_spinner=False)
def load_data(path: str, _signature) -> pd.DataFrame:
    """Load and shape the raw tracker sheet. `_signature` forces a reload
    whenever the file's mtime/size changes; it is not used otherwise."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET_NAME]

    rows = []
    for r in range(2, ws.max_row + 1):
        batch = ws.cell(row=r, column=COL["Batch"]).value
        if batch in (None, ""):
            continue
        rows.append(
            {
                "Batch": batch,
                "Participant": ws.cell(row=r, column=COL["Name"]).value,
                "Department": ws.cell(row=r, column=COL["Department"]).value or "(Blank)",
                "ProjectTitle": ws.cell(row=r, column=COL["ProjectTitle"]).value,
                "ARS": ws.cell(row=r, column=COL["ARS"]).value or 0,
                "D": ws.cell(row=r, column=COL["D"]).value,
                "M": ws.cell(row=r, column=COL["M"]).value,
                "A": ws.cell(row=r, column=COL["A"]).value,
                "I": ws.cell(row=r, column=COL["I"]).value,
                "C": ws.cell(row=r, column=COL["C"]).value,
            }
        )

    df = pd.DataFrame(rows)

    def is_done(v):
        return 1 if (v == "Completed" or v == 1) else 0

    for phase in ["D", "M", "A", "I", "C"]:
        df[f"{phase}_done"] = df[phase].apply(is_done)

    df["Completion_%"] = (
        df[["D_done", "M_done", "A_done", "I_done", "C_done"]].sum(axis=1) / 5 * 100
    )

    return df


# ----------------------------------------------------------------------
# Sidebar — data source + filters
# ----------------------------------------------------------------------
st.sidebar.title("⚙️ Data Source")
default_path = "SSBB.xlsx"
file_path = st.sidebar.text_input("Path to Excel file", value=default_path)

auto_refresh_on = st.sidebar.checkbox(
    "Auto-refresh dashboard", value=True,
    help="Automatically re-check the Excel file for changes on a timer.",
)
interval = st.sidebar.slider("Refresh interval (seconds)", 3, 60, 5, disabled=not auto_refresh_on)

if auto_refresh_on:
    if HAS_AUTOREFRESH:
        st_autorefresh(interval=interval * 1000, key="data_refresh")
    else:
        st.sidebar.warning(
            "Install `streamlit-autorefresh` for automatic polling:\n\n"
            "`pip install streamlit-autorefresh`\n\n"
            "Until then, use the Refresh button below or reload the page."
        )

if st.sidebar.button("🔄 Refresh now"):
    st.cache_data.clear()
    st.rerun()

if not os.path.exists(file_path):
    st.error(f"File not found: `{file_path}`. Update the path in the sidebar.")
    st.stop()

signature = get_file_signature(file_path)
df = load_data(file_path, signature)

last_loaded = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(signature[0]))
st.sidebar.caption(f"Excel last modified: {last_loaded}")

st.sidebar.title("🔍 Filters")
batches = ["All"] + sorted(df["Batch"].dropna().unique().tolist(), key=str)
participants = ["All"] + sorted(df["Participant"].dropna().unique().tolist(), key=str)
departments = ["All"] + sorted(df["Department"].dropna().unique().tolist(), key=str)

sel_batch = st.sidebar.selectbox("Batch", batches)
sel_participant = st.sidebar.selectbox("Participant's Name", participants)
sel_department = st.sidebar.selectbox("Department", departments)

fdf = df.copy()
if sel_batch != "All":
    fdf = fdf[fdf["Batch"] == sel_batch]
if sel_participant != "All":
    fdf = fdf[fdf["Participant"] == sel_participant]
if sel_department != "All":
    fdf = fdf[fdf["Department"] == sel_department]

if fdf.empty:
    st.warning("No data matches the selected filters.")
    st.stop()


# ----------------------------------------------------------------------
# Header
# ----------------------------------------------------------------------
st.markdown(
    """
    <div style="background:#0d1a3c;padding:20px;border-radius:12px;border:2px solid #f4d35e;text-align:center;">
        <h1 style="color:#f4d35e;font-family:sans-serif;margin:0;">SSBB PROJECT DASHBOARD</h1>
    </div>
    """,
    unsafe_allow_html=True,
)
st.write("")

# ----------------------------------------------------------------------
# KPI row 1
# ----------------------------------------------------------------------
total_projects = len(fdf)
total_participants = fdf["Participant"].nunique()
total_departments = fdf["Department"].nunique()
total_ars = fdf["ARS"].sum()

k1, k2, k3, k4 = st.columns(4)
kpi_style = """
<div style="background:{bg};padding:18px;border-radius:10px;text-align:center;">
    <div style="font-size:32px;font-weight:800;color:#111;">{value}</div>
    <div style="font-size:14px;color:#222;">{label}</div>
</div>
"""
k1.markdown(kpi_style.format(bg="#f6cbb0", value=total_projects, label="Total Projects"), unsafe_allow_html=True)
k2.markdown(kpi_style.format(bg="#d9c6ec", value=total_participants, label="Participants"), unsafe_allow_html=True)
k3.markdown(kpi_style.format(bg="#f2b9c4", value=total_departments, label="Departments"), unsafe_allow_html=True)
k4.markdown(kpi_style.format(bg="#f7e07a", value=f"{total_ars:,.2f}", label="Total ARS"), unsafe_allow_html=True)

st.write("")

# ----------------------------------------------------------------------
# KPI row 2 — DMAIC phase counts
# ----------------------------------------------------------------------
phase_counts = {
    "Define": fdf["D_done"].sum(),
    "Measure": fdf["M_done"].sum(),
    "Analyze": fdf["A_done"].sum(),
    "Improve": fdf["I_done"].sum(),
    "Control": fdf["C_done"].sum(),
}
phase_colors = ["#f4788c", "#8fd93c", "#e08a3e", "#f2b8c6", "#f4e04d"]
p_cols = st.columns(5)
for col, (label, value), color in zip(p_cols, phase_counts.items(), phase_colors):
    col.markdown(kpi_style.format(bg=color, value=int(value), label=label), unsafe_allow_html=True)

st.write("")

# ----------------------------------------------------------------------
# Charts
# ----------------------------------------------------------------------
left, right = st.columns([1.1, 1])

with left:
    st.subheader("Participant Completion Status (%)")
    part_completion = (
        fdf.groupby("Participant")["Completion_%"].mean().sort_values(ascending=True)
    )
    fig = go.Figure(
        go.Bar(
            x=part_completion.values,
            y=part_completion.index,
            orientation="h",
            marker_color=[
                "#4CAF50" if v >= 80 else "#f2e35b" if v >= 60 else "#e08a3e"
                for v in part_completion.values
            ],
            text=[f"{v:.2f}" for v in part_completion.values],
            textposition="outside",
        )
    )
    fig.update_layout(
        height=max(400, 22 * len(part_completion)),
        margin=dict(l=10, r=10, t=10, b=10),
        xaxis=dict(range=[0, 110], title=None),
        yaxis=dict(title=None),
        plot_bgcolor="white",
    )
    st.plotly_chart(fig, use_container_width=True)

with right:
    st.subheader("Total ARS by Department")
    dept_ars = fdf.groupby("Department")["ARS"].sum().sort_values(ascending=True)
    fig2 = go.Figure(
        go.Bar(
            x=dept_ars.values,
            y=dept_ars.index,
            orientation="h",
            marker_color="#f2d34d",
            text=[f"{v:,.0f}" for v in dept_ars.values],
            textposition="outside",
        )
    )
    fig2.update_layout(
        height=max(320, 34 * len(dept_ars)),
        margin=dict(l=10, r=10, t=10, b=10),
        xaxis=dict(title=None),
        yaxis=dict(title=None),
        plot_bgcolor="white",
    )
    st.plotly_chart(fig2, use_container_width=True)

    st.subheader("Overall Project Completion (%)")
    overall_completion = fdf["Completion_%"].mean()
    gauge = go.Figure(
        go.Indicator(
            mode="gauge+number",
            value=overall_completion,
            number={"suffix": "", "font": {"size": 40}},
            gauge={
                "axis": {"range": [0, 100]},
                "bar": {"color": "#4CAF50"},
                "bgcolor": "white",
                "steps": [{"range": [0, 100], "color": "#eeeeee"}],
                "threshold": {
                    "line": {"color": "#f4e04d", "width": 4},
                    "thickness": 0.9,
                    "value": 80,
                },
            },
        )
    )
    gauge.update_layout(height=280, margin=dict(l=20, r=20, t=10, b=10))
    st.plotly_chart(gauge, use_container_width=True)

st.caption(
    "Dashboard auto-refreshes from the Excel file on disk — edit and save "
    "`SSBB.xlsx` and the numbers above will update on the next refresh."
)
